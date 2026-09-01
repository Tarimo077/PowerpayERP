from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
import base64
import tempfile
import jwt
from django.contrib.auth.models import User
from django.core import mail
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from .models import (
    AuditLog,
    ChatInvitation,
    ChatMembership,
    ChatMessage,
    ChatPresence,
    ChatThread,
    Document,
    ItemRequest,
    LeaveAllocation,
    LeaveRequest,
    Notification,
    Organization,
    OutboundEmail,
    Profile,
    Task,
    Timesheet,
    TimesheetEntry,
    EmailOTP,
    PaymentVoucher,
    PaymentVoucherReceipt,
    UserInvite,
)


class TenantIsolationTests(TestCase):
    def setUp(self):
        self.o1 = Organization.objects.create(
            name="One", slug="one", business_email="one@test.com"
        )
        self.o2 = Organization.objects.create(
            name="Two", slug="two", business_email="two@test.com"
        )
        self.u1 = User.objects.create_user("one", password="testpass123")
        self.p1 = Profile.objects.create(
            user=self.u1, organization=self.o1, role="employee"
        )
        self.u2 = User.objects.create_user("two", password="testpass123")
        self.p2 = Profile.objects.create(
            user=self.u2, organization=self.o2, role="employee"
        )
        self.task = Task.objects.create(
            organization=self.o2,
            title="Secret",
            assigned_to=self.p2,
            created_by=self.u2,
            due_date=timezone.localdate() + timedelta(days=1),
        )

    def api_authorization(self, user, password="testpass123"):
        response = self.client.post(
            "/api/token/", {"email": user.email, "password": password}
        )
        self.assertEqual(response.status_code, 200, response.content)
        return f"Bearer {response.json()['access']}"

    def test_other_tenant_task_is_not_visible(self):
        self.client.login(username="one", password="testpass123")
        response = self.client.get(reverse("task_detail", args=[self.task.pk]))
        self.assertEqual(response.status_code, 404)

    def test_employee_profile_shows_information_and_only_updates_name(self):
        self.u1.first_name = "Original"
        self.u1.last_name = "Employee"
        self.u1.email = "original@example.com"
        self.u1.save(update_fields=["first_name", "last_name", "email"])
        self.p1.employee_id = "EMP-0042"
        self.p1.position = "Operations Officer"
        self.p1.phone = "+254700000000"
        self.p1.save(update_fields=["employee_id", "position", "phone"])

        self.client.force_login(self.u1)
        page = self.client.get(reverse("profile"))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "EMP-0042")
        self.assertContains(page, "Operations Officer")
        self.assertContains(page, "original@example.com")
        self.assertContains(page, "Edit name")

        response = self.client.post(
            reverse("profile"),
            {
                "first_name": "Updated",
                "last_name": "Person",
                "email": "attacker@example.com",
                "role": "admin",
                "employee_id": "ADMIN-9999",
            },
        )
        self.assertRedirects(response, reverse("profile"))
        self.u1.refresh_from_db()
        self.p1.refresh_from_db()
        self.assertEqual(self.u1.first_name, "Updated")
        self.assertEqual(self.u1.last_name, "Person")
        self.assertEqual(self.u1.email, "original@example.com")
        self.assertEqual(self.p1.role, "employee")
        self.assertEqual(self.p1.employee_id, "EMP-0042")
        self.assertTrue(
            AuditLog.objects.filter(
                actor=self.u1, action="profile_name_updated"
            ).exists()
        )

    def test_private_document_download_is_authorized(self):
        other_user = User.objects.create_user("document-owner", password="testpass123")
        other_profile = Profile.objects.create(
            user=other_user,
            organization=self.o1,
            role="employee",
        )
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            document = Document.objects.create(
                organization=self.o1,
                title="Private plan",
                file=SimpleUploadedFile("plan.txt", b"confidential"),
                visibility="private",
                owner=other_profile,
                uploaded_by=other_user,
            )
            self.client.force_login(self.u1)
            self.assertEqual(
                self.client.get(
                    reverse("document_download", args=[document.pk])
                ).status_code,
                404,
            )
            self.client.force_login(other_user)
            response = self.client.get(reverse("document_download", args=[document.pk]))
            self.assertEqual(response.status_code, 200)
            response.close()

    def test_api_requires_highest_organization_role_and_is_tenant_scoped(self):
        self.client.force_login(self.u1)
        self.assertEqual(self.client.get("/api/tasks/").status_code, 401)
        self.assertEqual(self.client.get("/api/docs/").status_code, 403)

        admin = User.objects.create_user(
            "org-admin", email="org-admin@example.com", password="testpass123"
        )
        Profile.objects.create(user=admin, organization=self.o1, role="admin")
        self.client.force_login(admin)
        self.assertEqual(self.client.get("/api/tasks/").status_code, 401)
        self.assertEqual(self.client.get("/api/docs/").status_code, 200)
        self.assertEqual(self.client.get("/api/schema/").status_code, 200)

        authorization = self.api_authorization(admin)
        response = self.client.get(
            "/api/tasks/", HTTP_AUTHORIZATION=authorization
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["count"], 0)

    def test_api_rejects_cross_tenant_relations_and_generic_status_changes(self):
        admin = User.objects.create_user(
            "api-admin", email="api-admin@example.com", password="testpass123"
        )
        Profile.objects.create(user=admin, organization=self.o1, role="admin")
        own_task = Task.objects.create(
            organization=self.o1,
            title="Own task",
            assigned_to=self.p1,
            created_by=self.u1,
            due_date=timezone.localdate() + timedelta(days=2),
        )
        authorization = self.api_authorization(admin)
        cross_tenant = self.client.post(
            "/api/tasks/",
            {
                "title": "Invalid API assignment",
                "organization": self.o2.pk,
                "assigned_to": self.p2.pk,
                "priority": "medium",
                "start_date": str(timezone.localdate()),
                "due_date": str(timezone.localdate() + timedelta(days=2)),
            },
            HTTP_AUTHORIZATION=authorization,
        )
        self.assertEqual(cross_tenant.status_code, 400)
        changed = self.client.patch(
            f"/api/tasks/{own_task.pk}/",
            {"status": "completed"},
            content_type="application/json",
            HTTP_AUTHORIZATION=authorization,
        )
        self.assertEqual(changed.status_code, 200)
        own_task.refresh_from_db()
        self.assertEqual(own_task.status, "assigned")

    def test_platform_admin_api_is_global_and_filterable(self):
        platform_admin = User.objects.create_superuser(
            "platform-api", "platform@example.com", "testpass123"
        )
        own_task = Task.objects.create(
            organization=self.o1,
            title="Visible globally",
            assigned_to=self.p1,
            created_by=self.u1,
            priority="urgent",
            due_date=timezone.localdate() + timedelta(days=2),
        )
        authorization = self.api_authorization(platform_admin)
        response = self.client.get(
            "/api/tasks/",
            {"organization": self.o1.pk, "priority": "urgent"},
            HTTP_AUTHORIZATION=authorization,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["count"], 1)
        self.assertEqual(response.json()["results"][0]["id"], own_task.pk)

        created = self.client.post(
            "/api/tasks/",
            {
                "organization": self.o2.pk,
                "title": "Platform-created task",
                "assigned_to": self.p2.pk,
                "priority": "medium",
                "start_date": str(timezone.localdate()),
                "due_date": str(timezone.localdate() + timedelta(days=3)),
            },
            HTTP_AUTHORIZATION=authorization,
        )
        self.assertEqual(created.status_code, 201, created.content)
        self.assertEqual(created.json()["organization"], self.o2.pk)

    def test_jwt_tokens_are_admin_only_and_use_reference_lifetimes(self):
        self.u1.email = "employee-token@example.com"
        self.u1.save(update_fields=["email"])
        denied = self.client.post(
            "/api/token/",
            {"email": self.u1.email, "password": "testpass123"},
        )
        self.assertEqual(denied.status_code, 401)

        admin = User.objects.create_user(
            "token-admin",
            email="token-admin@example.com",
            password="testpass123",
        )
        Profile.objects.create(user=admin, organization=self.o1, role="admin")
        issued = self.client.post(
            "/api/token/",
            {"email": admin.email, "password": "testpass123"},
        )
        self.assertEqual(issued.status_code, 200, issued.content)
        tokens = issued.json()
        access_claims = jwt.decode(
            tokens["access"], options={"verify_signature": False}
        )
        refresh_claims = jwt.decode(
            tokens["refresh"], options={"verify_signature": False}
        )
        self.assertEqual(access_claims["exp"] - access_claims["iat"], 24 * 60 * 60)
        self.assertEqual(
            refresh_claims["exp"] - refresh_claims["iat"], 3 * 24 * 60 * 60
        )
        self.assertEqual(access_claims["organization_id"], self.o1.pk)
        self.assertEqual(tokens["access_expires_in"], 24 * 60 * 60)
        self.assertEqual(tokens["refresh_expires_in"], 3 * 24 * 60 * 60)

        bearer_response = self.client.get(
            "/api/tasks/", HTTP_AUTHORIZATION=f"Bearer {tokens['access']}"
        )
        self.assertEqual(bearer_response.status_code, 200)

        rotated = self.client.post(
            "/api/token/refresh/", {"refresh": tokens["refresh"]}
        )
        self.assertEqual(rotated.status_code, 200, rotated.content)
        self.assertIn("access", rotated.json())
        self.assertIn("refresh", rotated.json())
        reused = self.client.post("/api/token/refresh/", {"refresh": tokens["refresh"]})
        self.assertEqual(reused.status_code, 401)

    def test_email_login_requires_otp_before_session_login(self):
        self.u1.email = "one@example.com"
        self.u1.save(update_fields=["email"])
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(
                reverse("login"),
                {"email": "one@example.com", "password": "testpass123"},
            )
        self.assertRedirects(response, reverse("verify_otp"))
        self.assertNotIn("_auth_user_id", self.client.session)
        self.assertTrue(EmailOTP.objects.filter(user=self.u1).exists())
        self.assertTrue(
            OutboundEmail.objects.filter(
                recipient="one@example.com", status="sent"
            ).exists()
        )

    def test_login_rate_limit_blocks_repeated_password_attempts(self):
        cache.clear()
        self.u1.email = "limited@example.com"
        self.u1.save(update_fields=["email"])
        for _ in range(5):
            response = self.client.post(
                reverse("login"),
                {"email": "limited@example.com", "password": "wrong-password"},
            )
            self.assertEqual(response.status_code, 200)
        blocked = self.client.post(
            reverse("login"),
            {"email": "limited@example.com", "password": "wrong-password"},
        )
        self.assertEqual(blocked.status_code, 429)

    def test_employee_invite_gets_generated_id_and_valid_activation_page(self):
        invite = UserInvite.objects.create(
            email="new.employee@example.com",
            organization=self.o1,
            role="employee",
            invited_by=self.u1,
        )
        self.assertRegex(invite.employee_id, r"^EMP-\d{4}$")
        response = self.client.get(reverse("accept_invite", args=[invite.token]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Activate account")
        self.assertContains(response, "at least 8 characters")
        self.assertContains(response, "installPasswordToggles")

    def test_admin_add_employee_creates_invite_without_credentials(self):
        self.p1.role = "admin"
        self.p1.save(update_fields=["role"])
        self.client.login(username="one", password="testpass123")
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(
                reverse("employee_create"),
                {
                    "email": "invited@example.com",
                    "role": "employee",
                    "position": "Analyst",
                    "department": "",
                    "manager": self.p1.pk,
                },
            )
        self.assertRedirects(response, reverse("employees"))
        invite = UserInvite.objects.get(email="invited@example.com")
        self.assertRegex(invite.employee_id, r"^EMP-\d{4}$")
        self.assertFalse(User.objects.filter(email="invited@example.com").exists())

    def test_create_and_add_another_returns_fresh_employee_form(self):
        self.p1.role = "admin"
        self.p1.save(update_fields=["role"])
        self.client.login(username="one", password="testpass123")
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(
                reverse("employee_create"),
                {
                    "email": "another@example.com",
                    "role": "employee",
                    "position": "Analyst",
                    "department": "",
                    "manager": self.p1.pk,
                    "_add_another": "1",
                },
            )
        self.assertRedirects(response, reverse("employee_create"))
        self.assertTrue(UserInvite.objects.filter(email="another@example.com").exists())

    def test_storyboard_is_available_to_authenticated_users(self):
        self.client.login(username="one", password="testpass123")
        response = self.client.get(reverse("storyboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Learn PowerpayERP one workflow at a time")
        self.assertContains(response, "Roles and data boundaries")

    def test_notification_dropdown_and_read_actions(self):
        unread = Notification.objects.create(
            user=self.u1,
            title="New task assigned",
            message="Prepare the monthly report.",
            url="/tasks/",
        )
        Notification.objects.create(
            user=self.u1,
            title="Previous update",
            message="Already reviewed.",
            is_read=True,
        )
        other = Notification.objects.create(
            user=self.u2,
            title="Private update",
            message="Another tenant's notification.",
        )
        self.client.force_login(self.u1)

        response = self.client.get(reverse("notifications"))
        self.assertContains(response, 'id="notification-panel"')
        self.assertContains(response, "New task assigned")
        self.assertContains(response, "1 unread update")
        unread.refresh_from_db()
        self.assertFalse(unread.is_read)

        result = self.client.post(reverse("notification_read", args=[unread.pk]))
        self.assertEqual(result.status_code, 200)
        self.assertEqual(result.json()["unread"], 0)
        unread.refresh_from_db()
        self.assertTrue(unread.is_read)
        self.assertEqual(
            self.client.post(reverse("notification_read", args=[other.pk])).status_code,
            404,
        )

    def test_mark_all_notifications_read_is_scoped_to_current_user(self):
        mine = Notification.objects.create(user=self.u1, title="My update")
        other = Notification.objects.create(user=self.u2, title="Other update")
        self.client.force_login(self.u1)
        result = self.client.post(reverse("notifications_read_all"))
        self.assertEqual(result.status_code, 200)
        self.assertEqual(result.json()["unread"], 0)
        mine.refresh_from_db()
        other.refresh_from_db()
        self.assertTrue(mine.is_read)
        self.assertFalse(other.is_read)

    def test_superuser_can_use_platform_control_pages(self):
        root = User.objects.create_superuser("root", "root@example.com", "adminpass123")
        self.client.force_login(root)
        for url in [
            reverse("platform_organizations"),
            reverse("platform_organization_detail", args=[self.o1.pk]),
            reverse("platform_users"),
            reverse("platform_activity"),
        ]:
            self.assertEqual(self.client.get(url).status_code, 200)

    def test_organization_user_cannot_use_platform_control_pages(self):
        self.client.force_login(self.u1)
        response = self.client.get(reverse("platform_organizations"))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("login"), response.url)

    def test_suspending_an_organization_ends_member_access(self):
        root = User.objects.create_superuser("root", "root@example.com", "adminpass123")
        self.client.force_login(root)
        self.client.post(reverse("platform_organization_status", args=[self.o1.pk]))
        self.o1.refresh_from_db()
        self.assertFalse(self.o1.is_active)

        self.client.force_login(self.u1)
        response = self.client.get(reverse("dashboard"))
        self.assertRedirects(response, reverse("login"))
        self.assertNotIn("_auth_user_id", self.client.session)

    def test_employee_can_create_route_and_export_itemized_payment_voucher(self):
        self.client.force_login(self.u1)
        manager_user = User.objects.create_user(
            "manager", email="manager@example.com", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(
                reverse("payment_voucher_create"),
                {
                    "date": str(timezone.localdate()),
                    "payee": "Acme Supplies",
                    "payee_id_number": "P12345",
                    "department": "",
                    "lines-TOTAL_FORMS": "2",
                    "lines-INITIAL_FORMS": "0",
                    "lines-MIN_NUM_FORMS": "1",
                    "lines-MAX_NUM_FORMS": "1000",
                    "lines-0-particulars": "Office stationery",
                    "lines-0-amount": "1500.00",
                    "lines-1-particulars": "Delivery",
                    "lines-1-amount": "250.00",
                },
            )
        voucher = PaymentVoucher.objects.get(organization=self.o1)
        self.assertRedirects(
            response, reverse("payment_voucher_detail", args=[voucher.pk])
        )
        self.assertRegex(voucher.number, r"^PV-\d{4}-0001$")
        self.assertEqual(voucher.total_amount, Decimal("1750.00"))
        self.assertEqual(voucher.status, "submitted")
        self.assertEqual(voucher.requested_approver, manager)
        self.assertEqual(voucher.approval_task.assigned_to, manager)
        self.assertEqual(voucher.approval_task.priority, "medium")
        self.assertIn("Approve payment voucher", voucher.approval_task.title)
        self.assertEqual(len(mail.outbox), 1)
        pdf = self.client.get(reverse("payment_voucher_pdf", args=[voucher.pk]))
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF"))

    def test_payment_voucher_accepts_multiple_pdf_receipts_for_manager_review(self):
        manager_user = User.objects.create_user(
            "receipt-manager", email="manager@example.com", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        self.client.force_login(self.u1)
        data = {
            "date": str(timezone.localdate()),
            "payee": "Receipt Supplier",
            "payee_id_number": "SUP-1",
            "department": "",
            "lines-TOTAL_FORMS": "1",
            "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "1",
            "lines-MAX_NUM_FORMS": "1000",
            "lines-0-particulars": "Supplies",
            "lines-0-amount": "500.00",
            "receipts": [
                SimpleUploadedFile(
                    "receipt-one.pdf",
                    b"%PDF-1.4\nreceipt one",
                    content_type="application/pdf",
                ),
                SimpleUploadedFile(
                    "receipt-two.pdf",
                    b"%PDF-1.4\nreceipt two",
                    content_type="application/pdf",
                ),
            ],
        }
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root,
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        ):
            response = self.client.post(reverse("payment_voucher_create"), data)
            voucher = PaymentVoucher.objects.get(payee="Receipt Supplier")
            self.assertRedirects(
                response, reverse("payment_voucher_detail", args=[voucher.pk])
            )
            self.assertEqual(voucher.receipts.count(), 2)
            receipt = voucher.receipts.get(original_name="receipt-one.pdf")
            added = self.client.post(
                reverse("payment_voucher_receipt_add", args=[voucher.pk]),
                {
                    "receipts": SimpleUploadedFile(
                        "receipt-three.pdf",
                        b"%PDF-1.4\nreceipt three",
                        content_type="application/pdf",
                    )
                },
            )
            self.assertRedirects(
                added, reverse("payment_voucher_detail", args=[voucher.pk])
            )
            self.assertEqual(voucher.receipts.count(), 3)
            self.client.force_login(manager_user)
            review = self.client.get(
                reverse("payment_voucher_detail", args=[voucher.pk])
            )
            self.assertContains(review, "receipt-one.pdf")
            self.assertContains(review, "receipt-two.pdf")
            self.assertContains(review, "receipt-three.pdf")
            opened = self.client.get(
                reverse("payment_voucher_receipt", args=[voucher.pk, receipt.pk])
            )
            self.assertEqual(opened.status_code, 200)
            self.assertEqual(opened["Content-Type"], "application/pdf")
            opened.close()
            self.client.force_login(self.u2)
            self.assertEqual(
                self.client.get(
                    reverse("payment_voucher_receipt", args=[voucher.pk, receipt.pk])
                ).status_code,
                404,
            )

    def test_payment_voucher_rejects_a_fake_pdf_receipt(self):
        manager_user = User.objects.create_user(
            "receipt-manager", email="manager@example.com", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        self.client.force_login(self.u1)
        data = {
            "date": str(timezone.localdate()),
            "payee": "Invalid Receipt",
            "payee_id_number": "",
            "department": "",
            "lines-TOTAL_FORMS": "1",
            "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "1",
            "lines-MAX_NUM_FORMS": "1000",
            "lines-0-particulars": "Supplies",
            "lines-0-amount": "500.00",
            "receipts": SimpleUploadedFile(
                "fake.pdf", b"this is not a pdf", content_type="application/pdf"
            ),
        }
        response = self.client.post(reverse("payment_voucher_create"), data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "file contents are not a valid PDF")
        self.assertFalse(
            PaymentVoucher.objects.filter(payee="Invalid Receipt").exists()
        )

    def test_blank_voucher_line_amount_returns_form_error_instead_of_crashing(self):
        manager_user = User.objects.create_user(
            "manager", email="manager@example.com", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        self.client.force_login(self.u1)
        response = self.client.post(
            reverse("payment_voucher_create"),
            {
                "date": str(timezone.localdate()),
                "payee": "Acme Supplies",
                "payee_id_number": "P12345",
                "department": "",
                "lines-TOTAL_FORMS": "1",
                "lines-INITIAL_FORMS": "0",
                "lines-MIN_NUM_FORMS": "1",
                "lines-MAX_NUM_FORMS": "1000",
                "lines-0-particulars": "Office stationery",
                "lines-0-amount": "",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "This field is required")
        self.assertFalse(PaymentVoucher.objects.filter(organization=self.o1).exists())

    def test_payment_voucher_workflow_and_tenant_boundary(self):
        self.p1.role = "admin"
        self.p1.save(update_fields=["role"])
        manager_user = User.objects.create_user("manager", password="testpass123")
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        senior_user = User.objects.create_user(
            "senior-manager", email="senior@example.com", password="testpass123"
        )
        senior = Profile.objects.create(
            user=senior_user, organization=self.o1, role="manager"
        )
        manager.manager = senior
        manager.save(update_fields=["manager"])
        review_task = Task.objects.create(
            organization=self.o1,
            title="Approve voucher",
            assigned_to=manager,
            created_by=self.u1,
            priority="high",
            status="assigned",
            start_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=2),
        )
        voucher = PaymentVoucher.objects.create(
            organization=self.o1,
            number="PV-2026-0001",
            date=timezone.localdate(),
            payee="Payee",
            prepared_by=self.u1,
            requested_approver=manager,
            approval_task=review_task,
            status="submitted",
        )
        other_manager_user = User.objects.create_user(
            "other-manager", password="testpass123"
        )
        Profile.objects.create(
            user=other_manager_user, organization=self.o1, role="manager"
        )
        self.client.force_login(other_manager_user)
        self.client.post(
            reverse("payment_voucher_action", args=[voucher.pk, "approve"]),
            {"notes": "Not mine"},
        )
        voucher.refresh_from_db()
        self.assertEqual(voucher.status, "submitted")
        self.client.force_login(manager_user)
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            self.client.post(
                reverse("payment_voucher_action", args=[voucher.pk, "escalate"]),
                {"notes": "Senior review required"},
            )
        voucher.refresh_from_db()
        review_task.refresh_from_db()
        self.assertEqual(voucher.requested_approver, senior)
        self.assertEqual(review_task.assigned_to, senior)
        self.assertEqual(mail.outbox[-1].to, ["senior@example.com"])
        self.client.force_login(senior_user)
        self.client.post(
            reverse("payment_voucher_action", args=[voucher.pk, "approve"]),
            {"notes": "Approved"},
        )
        self.client.force_login(self.u1)
        self.client.post(
            reverse("payment_voucher_action", args=[voucher.pk, "paid"]),
            {"payment_received_by": "Jane Doe"},
        )
        voucher.refresh_from_db()
        self.assertEqual(voucher.status, "paid")
        self.assertEqual(voucher.payment_received_by, "Jane Doe")
        self.client.force_login(self.u2)
        self.p2.role = "manager"
        self.p2.save(update_fields=["role"])
        self.assertEqual(
            self.client.get(
                reverse("payment_voucher_detail", args=[voucher.pk])
            ).status_code,
            404,
        )

    def test_employee_can_assign_task_only_to_self_or_manager_and_email_is_sent(self):
        self.u1.email = "employee@example.com"
        self.u1.save(update_fields=["email"])
        manager_user = User.objects.create_user(
            "manager", email="manager@example.com", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        self.client.force_login(self.u1)
        data = {
            "title": "Prepare report",
            "description": "Monthly report",
            "instructions": "",
            "assigned_to": manager.pk,
            "department": "",
            "priority": "high",
            "start_date": str(timezone.localdate()),
            "due_date": str(timezone.localdate() + timedelta(days=2)),
        }
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(reverse("task_create"), data)
        task = Task.objects.get(title="Prepare report")
        self.assertRedirects(response, reverse("tasks"))
        self.assertEqual(task.assigned_to, manager)
        self.assertEqual(task.status, "assigned")
        self.assertEqual(mail.outbox[-1].to, ["manager@example.com"])
        self.assertIn("New task assigned", mail.outbox[-1].subject)

        data.update({"title": "Invalid assignment", "assigned_to": self.p2.pk})
        response = self.client.post(reverse("task_create"), data)
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Task.objects.filter(title="Invalid assignment").exists())

    def test_assignee_can_change_task_status_from_popup(self):
        task = Task.objects.create(
            organization=self.o1,
            title="Prepare minutes",
            assigned_to=self.p1,
            created_by=self.u1,
            status="assigned",
            due_date=timezone.localdate() + timedelta(days=2),
        )
        self.client.force_login(self.u1)
        page = self.client.get(reverse("tasks"))
        self.assertContains(page, "Change status")
        popup = self.client.get(reverse("task_status", args=[task.pk]))
        self.assertEqual(popup.status_code, 200)
        self.assertContains(popup, "Current status")
        self.assertContains(popup, "Assigned · Not started")
        self.assertContains(popup, "Start / resume work")
        self.assertContains(popup, "Submit for review")
        response = self.client.post(
            reverse("task_status", args=[task.pk]),
            {"status": "in_progress", "note": "Work has started."},
        )
        self.assertRedirects(response, reverse("task_detail", args=[task.pk]))
        task.refresh_from_db()
        self.assertEqual(task.status, "in_progress")
        self.assertIsNotNone(task.actual_started_at)
        self.assertIsNone(task.actual_completed_at)

    def test_manager_changes_direct_report_task_status_and_employee_is_notified(self):
        manager_user = User.objects.create_user(
            "status-manager", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        task = Task.objects.create(
            organization=self.o1,
            title="Close the report",
            assigned_to=self.p1,
            created_by=manager_user,
            status="submitted",
            due_date=timezone.localdate() + timedelta(days=1),
        )
        self.client.force_login(manager_user)
        response = self.client.post(
            reverse("task_status", args=[task.pk]),
            {"status": "approved", "note": "Reviewed and approved."},
        )
        self.assertRedirects(response, reverse("task_detail", args=[task.pk]))
        task.refresh_from_db()
        self.assertEqual(task.status, "approved")
        self.assertIsNotNone(task.actual_started_at)
        self.assertIsNotNone(task.actual_completed_at)
        notice = Notification.objects.get(
            user=self.u1, title="Task status changed to Approved"
        )
        self.assertEqual(notice.message, "Reviewed and approved.")

    def test_manager_dashboard_has_hierarchy_task_calendar_with_status(self):
        manager_user = User.objects.create_user(
            "calendar-manager", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        lead_user = User.objects.create_user("calendar-lead", password="testpass123")
        lead = Profile.objects.create(
            user=lead_user, organization=self.o1, role="manager", manager=manager
        )
        employee_user = User.objects.create_user(
            "calendar-employee", password="testpass123"
        )
        employee = Profile.objects.create(
            user=employee_user, organization=self.o1, role="employee", manager=lead
        )
        unrelated_user = User.objects.create_user(
            "calendar-unrelated", password="testpass123"
        )
        unrelated = Profile.objects.create(
            user=unrelated_user, organization=self.o1, role="employee"
        )
        Task.objects.create(
            organization=self.o1,
            title="Lead planning",
            assigned_to=lead,
            created_by=manager_user,
            status="in_progress",
            priority="medium",
            due_date=timezone.localdate() + timedelta(days=2),
        )
        Task.objects.create(
            organization=self.o1,
            title="Employee field work",
            assigned_to=employee,
            created_by=lead_user,
            status="submitted",
            priority="high",
            due_date=timezone.localdate() + timedelta(days=3),
        )
        Task.objects.create(
            organization=self.o1,
            title="Manager follow-up",
            assigned_to=manager,
            created_by=manager_user,
            status="rejected",
            priority="low",
            due_date=timezone.localdate() + timedelta(days=1),
        )
        Task.objects.create(
            organization=self.o1,
            title="Unrelated work",
            assigned_to=unrelated,
            created_by=unrelated_user,
            status="assigned",
            due_date=timezone.localdate() + timedelta(days=1),
        )
        self.client.force_login(manager_user)
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Team task calendar")
        self.assertEqual(response.context["team_member_count"], 2)
        events = response.context["team_task_events"]
        self.assertEqual(
            {event["title"] for event in events},
            {"Lead planning", "Employee field work"},
        )
        self.assertEqual(
            {event["status_label"] for event in events}, {"In Progress", "Submitted"}
        )
        self.assertEqual(
            [
                (event["title"], event["status_label"])
                for event in response.context["task_events"]
            ],
            [("Manager follow-up", "Rejected")],
        )
        self.assertContains(response, 'id="team-task-calendar-events"')
        self.assertContains(response, "event.status_label")
        self.assertContains(response, 'data-calendar-tab="personal"')
        self.assertContains(response, 'data-calendar-tab="team"')
        self.assertContains(response, 'data-calendar-panel="team"')

    def test_employee_can_edit_and_delete_own_draft_timesheet(self):
        sheet = Timesheet.objects.create(
            organization=self.o1,
            employee=self.p1,
            period_start=date(2026, 8, 1),
            period_end=date(2026, 8, 31),
            status="draft",
            service_contract="Original",
        )
        entry = TimesheetEntry.objects.create(
            timesheet=sheet,
            date=date(2026, 8, 3),
            task_performed="Draft activity",
            hours=8,
            days_worked=1,
        )
        self.client.force_login(self.u1)
        response = self.client.post(
            reverse("timesheet_edit", args=[sheet.pk]),
            {
                "service_contract": "Updated contract",
                "financing": "Internal",
                "contract_number": "TS-22",
                "country": "Kenya",
                "place_of_assignment": "Nairobi",
                "initial_budget_days": "30",
            },
        )
        self.assertRedirects(response, reverse("timesheet_detail", args=[sheet.pk]))
        sheet.refresh_from_db()
        self.assertEqual(sheet.service_contract, "Updated contract")
        self.assertEqual(sheet.period_start, date(2026, 8, 1))
        response = self.client.post(reverse("timesheet_delete", args=[sheet.pk]))
        self.assertRedirects(response, reverse("timesheets"))
        self.assertFalse(Timesheet.objects.filter(pk=sheet.pk).exists())
        self.assertFalse(TimesheetEntry.objects.filter(pk=entry.pk).exists())

    def test_popup_clones_remove_source_hidden_class(self):
        self.client.force_login(self.u1)
        response = self.client.get(reverse("dashboard"))
        self.assertContains(response, "copy.classList.remove('hidden')")
        self.assertContains(response, "form.getAttribute('action') || location.href")
        self.assertNotContains(response, "fetch(form.action||location.href")

    def test_timesheet_year_filter_is_unique_and_includes_recorded_years(self):
        Timesheet.objects.create(
            organization=self.o1,
            employee=self.p1,
            period_start=date(2022, 4, 1),
            period_end=date(2022, 4, 30),
            status="draft",
        )
        Timesheet.objects.create(
            organization=self.o1,
            employee=self.p1,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            status="draft",
        )
        self.client.force_login(self.u1)
        response = self.client.get(reverse("timesheets"), {"year": "2026"})
        self.assertEqual(response.context["year_options"], [2022, 2025, 2026, 2027])
        self.assertEqual(response.context["year_options"].count(2026), 1)
        self.assertContains(
            response,
            '<option value="2026" selected>2026</option>',
            count=1,
            html=True,
        )

    def test_submitted_timesheet_cannot_be_edited_or_deleted(self):
        sheet = Timesheet.objects.create(
            organization=self.o1,
            employee=self.p1,
            period_start=date(2026, 8, 1),
            period_end=date(2026, 8, 31),
            status="submitted",
            service_contract="Locked",
        )
        self.client.force_login(self.u1)
        response = self.client.get(reverse("timesheet_edit", args=[sheet.pk]))
        self.assertRedirects(response, reverse("timesheet_detail", args=[sheet.pk]))
        response = self.client.post(reverse("timesheet_delete", args=[sheet.pk]))
        self.assertRedirects(response, reverse("timesheet_detail", args=[sheet.pk]))
        sheet.refresh_from_db()
        self.assertEqual(sheet.service_contract, "Locked")

    def test_manager_cannot_see_employee_drafts_but_can_see_submitted_timesheets(self):
        manager_user = User.objects.create_user(
            "visibility-manager", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        draft = Timesheet.objects.create(
            organization=self.o1,
            employee=self.p1,
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
            status="draft",
            service_contract="Private draft",
        )
        submitted = Timesheet.objects.create(
            organization=self.o1,
            employee=self.p1,
            period_start=date(2026, 7, 1),
            period_end=date(2026, 7, 31),
            status="submitted",
            requested_approver=manager,
            service_contract="Ready for review",
        )
        own_draft = Timesheet.objects.create(
            organization=self.o1,
            employee=manager,
            period_start=date(2026, 8, 1),
            period_end=date(2026, 8, 31),
            status="draft",
            service_contract="Manager own draft",
        )
        self.client.force_login(manager_user)
        response = self.client.get(reverse("timesheets"), {"year": "2026"})
        self.assertNotIn(draft, list(response.context["timesheets"]))
        self.assertIn(submitted, list(response.context["timesheets"]))
        self.assertIn(own_draft, list(response.context["timesheets"]))
        self.assertRedirects(
            self.client.get(reverse("timesheet_detail", args=[draft.pk])),
            reverse("timesheets"),
        )
        self.assertRedirects(
            self.client.get(reverse("timesheet_export", args=[draft.pk, "pdf"])),
            reverse("timesheets"),
        )
        self.assertEqual(
            self.client.get(
                reverse("timesheet_detail", args=[submitted.pk])
            ).status_code,
            200,
        )
        delete_response = self.client.post(reverse("timesheet_delete", args=[draft.pk]))
        self.assertRedirects(
            delete_response,
            reverse("timesheet_detail", args=[draft.pk]),
            fetch_redirect_response=False,
        )
        self.assertTrue(Timesheet.objects.filter(pk=draft.pk).exists())

    def test_new_timesheet_prefills_assigned_tasks_and_employee_can_edit_them(self):
        worked_at = timezone.make_aware(datetime(2026, 8, 3, 9, 0))
        assigned = Task.objects.create(
            organization=self.o1,
            title="Prepare payroll",
            description="Process the monthly payroll.",
            assigned_to=self.p1,
            created_by=self.u1,
            status="completed",
            start_date=date(2026, 8, 3),
            due_date=date(2026, 8, 7),
            actual_started_at=worked_at,
            actual_completed_at=worked_at,
        )
        outside = Task.objects.create(
            organization=self.o1,
            title="Later task",
            assigned_to=self.p1,
            created_by=self.u1,
            status="assigned",
            start_date=date(2026, 9, 1),
            due_date=date(2026, 9, 2),
        )
        manager_user = User.objects.create_user(
            "timesheet-manager", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        self.client.force_login(self.u1)
        create_page = self.client.get(reverse("timesheet_create"))
        self.assertEqual(
            create_page.context["form"].fields["place_of_assignment"].initial, "Kenya"
        )
        response = self.client.post(
            reverse("timesheet_create"),
            {
                "month": "8",
                "year": "2026",
                "service_contract": "Digital energy services",
                "financing": "Internal",
                "contract_number": "TS-001",
                "country": "Kenya",
                "place_of_assignment": "Nairobi",
                "initial_budget_days": "120",
            },
        )
        sheet = Timesheet.objects.get(employee=self.p1)
        self.assertRedirects(response, reverse("timesheet_detail", args=[sheet.pk]))
        self.assertEqual(sheet.period_start, date(2026, 8, 1))
        self.assertEqual(sheet.period_end, date(2026, 8, 31))
        self.assertEqual(sheet.entries.count(), 1)
        entry = sheet.entries.get()
        self.assertEqual(entry.task, assigned)
        self.assertEqual(
            entry.task_performed, "Prepare payroll: Process the monthly payroll."
        )
        self.assertEqual(entry.hours, Decimal("8"))
        self.assertEqual(entry.days_worked, Decimal("1"))
        self.assertEqual(entry.location, "Kenya")
        self.assertFalse(sheet.entries.filter(task=outside).exists())

        blocked = self.client.post(
            reverse("timesheet_action", args=[sheet.pk]), {"action": "submit"}
        )
        self.assertRedirects(blocked, reverse("timesheet_detail", args=[sheet.pk]))
        sheet.refresh_from_db()
        self.assertEqual(sheet.status, "draft")
        response = self.client.post(
            reverse("timesheet_entry_edit", args=[sheet.pk, entry.pk]),
            {
                "date": "2026-08-04",
                "task": assigned.pk,
                "task_performed": "Payroll preparation and checks",
                "hours": "6.50",
                "days_worked": "0.75",
                "location": "Kenya, Nairobi",
                "description": "Updated work details",
                "notes": "",
                "supporting_document": "",
            },
        )
        self.assertRedirects(response, reverse("timesheet_detail", args=[sheet.pk]))
        entry.refresh_from_db()
        self.assertEqual(entry.hours, Decimal("6.50"))
        self.assertEqual(entry.task_performed, "Payroll preparation and checks")
        png = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
        )
        with tempfile.TemporaryDirectory() as media_root, self.settings(
            MEDIA_ROOT=media_root
        ):
            signed = self.client.post(
                reverse("timesheet_signature_upload", args=[sheet.pk]),
                {
                    "expert_signature": SimpleUploadedFile(
                        "expert.png", png, content_type="image/png"
                    )
                },
            )
            self.assertRedirects(signed, reverse("timesheet_detail", args=[sheet.pk]))
            annual = self.client.post(
                reverse("timesheet_export_year", args=[sheet.pk]),
                {"months": [str(month) for month in range(1, 13)]},
            )
            self.assertEqual(annual.status_code, 200)
            from openpyxl import load_workbook

            workbook = load_workbook(BytesIO(annual.content), data_only=False)
            self.assertEqual(len(workbook.sheetnames), 12)
            self.assertEqual(workbook.sheetnames[0], "TS - January 26")
            self.assertEqual(workbook.sheetnames[-1], "TS - December 26")
            august = workbook["TS - August 26"]
            self.assertEqual(august["A1"].value, "TIME SHEET FOR EXPERTS")
            self.assertEqual(august["C46"].value, "=SUM(C14:C44)")
            self.assertIn("Payroll preparation", august["H17"].value)
            self.assertFalse(any(image.width == 150 for image in august._images))
            self.assertEqual(august["A13"].fill.fgColor.rgb, "004B5563")
            self.assertEqual(august.sheet_properties.tabColor, None)
            from openpyxl.cell.cell import MergedCell

            for row in august.iter_rows(min_row=1, max_row=62, min_col=1, max_col=8):
                for cell in row:
                    if not isinstance(cell, MergedCell):
                        self.assertTrue(
                            all(
                                getattr(cell.border, side).style
                                for side in ["left", "right", "top", "bottom"]
                            )
                        )
            self.client.post(
                reverse("timesheet_action", args=[sheet.pk]), {"action": "submit"}
            )
            sheet.refresh_from_db()
            self.assertEqual(sheet.status, "submitted")
            self.client.force_login(manager_user)
            reviewed = self.client.post(
                reverse("timesheet_review", args=[sheet.pk]),
                {
                    "decision": "approved",
                    "notes": "Approved monthly record.",
                    "manager_signature": SimpleUploadedFile(
                        "manager.png", png, content_type="image/png"
                    ),
                    "consultant_name": "Jane Consultant",
                    "consultant_signature": SimpleUploadedFile(
                        "consultant.png", png, content_type="image/png"
                    ),
                },
            )
            self.assertRedirects(reviewed, reverse("timesheet_detail", args=[sheet.pk]))
            sheet.refresh_from_db()
            self.assertEqual(sheet.status, "approved")
            self.assertTrue(bool(sheet.manager_signature))
            self.assertTrue(bool(sheet.consultant_signature))

    def test_timesheet_export_selects_months_and_manual_work_needs_no_system_task(self):
        sheet = Timesheet.objects.create(
            organization=self.o1,
            employee=self.p1,
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
        )
        self.client.force_login(self.u1)
        added = self.client.post(
            reverse("timesheet_detail", args=[sheet.pk]),
            {
                "date": "2026-03-12",
                "task": "",
                "task_performed": "Unplanned stakeholder workshop",
                "hours": "4",
                "days_worked": "0.5",
                "location": "Nairobi",
                "description": "Work performed outside the task register.",
                "notes": "",
            },
        )
        self.assertRedirects(added, reverse("timesheet_detail", args=[sheet.pk]))
        entry = sheet.entries.get()
        self.assertIsNone(entry.task)
        self.assertEqual(entry.task_performed, "Unplanned stakeholder workshop")
        chooser = self.client.get(reverse("timesheet_export_year", args=[sheet.pk]))
        self.assertEqual(chooser.status_code, 200)
        self.assertContains(chooser, "Choose months to export")
        exported = self.client.post(
            reverse("timesheet_export_year", args=[sheet.pk]), {"months": ["3", "6"]}
        )
        self.assertEqual(exported.status_code, 200)
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(exported.content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["TS - March 26", "TS - June 26"])
        self.assertIn(
            "Unplanned stakeholder workshop", workbook["TS - March 26"]["H25"].value
        )

    def test_timesheet_prefill_uses_actual_task_work_dates(self):
        started = timezone.make_aware(datetime(2026, 8, 7, 9, 0))
        completed = timezone.make_aware(datetime(2026, 8, 11, 17, 0))
        task = Task.objects.create(
            organization=self.o1,
            title="Field verification",
            description="Verify installation sites.",
            assigned_to=self.p1,
            created_by=self.u1,
            status="completed",
            start_date=date(2026, 8, 1),
            due_date=date(2026, 8, 20),
            actual_started_at=started,
            actual_completed_at=completed,
        )
        self.client.force_login(self.u1)
        response = self.client.post(
            reverse("timesheet_create"),
            {
                "month": "8",
                "year": "2026",
                "service_contract": "Field work",
                "financing": "Internal",
                "contract_number": "TS-DATES",
                "country": "Kenya",
                "place_of_assignment": "Kenya",
                "initial_budget_days": "30",
            },
        )
        sheet = Timesheet.objects.get(employee=self.p1, period_start=date(2026, 8, 1))
        self.assertRedirects(response, reverse("timesheet_detail", args=[sheet.pk]))
        self.assertEqual(
            list(
                sheet.entries.filter(task=task)
                .order_by("date")
                .values_list("date", flat=True)
            ),
            [date(2026, 8, 7), date(2026, 8, 10), date(2026, 8, 11)],
        )
        self.assertTrue(
            all(
                entry.task_performed == "Field verification: Verify installation sites."
                for entry in sheet.entries.filter(task=task)
            )
        )

    def test_employee_can_remove_a_prefilled_timesheet_task(self):
        task = Task.objects.create(
            organization=self.o1,
            title="Removable task",
            assigned_to=self.p1,
            created_by=self.u1,
            status="assigned",
            start_date=date(2026, 8, 3),
            due_date=date(2026, 8, 4),
        )
        sheet = Timesheet.objects.create(
            organization=self.o1,
            employee=self.p1,
            period_start=date(2026, 8, 3),
            period_end=date(2026, 8, 9),
        )
        entry = TimesheetEntry.objects.create(
            timesheet=sheet,
            date=date(2026, 8, 3),
            task=task,
            task_performed=task.title,
            hours=0,
        )
        self.client.force_login(self.u1)
        response = self.client.post(
            reverse("timesheet_entry_delete", args=[sheet.pk, entry.pk])
        )
        self.assertRedirects(response, reverse("timesheet_detail", args=[sheet.pk]))
        self.assertFalse(TimesheetEntry.objects.filter(pk=entry.pk).exists())

    def test_timesheet_moves_up_manager_chain_and_managers_can_edit(self):
        senior_user = User.objects.create_user(
            "senior-manager", email="senior@example.com", password="testpass123"
        )
        senior = Profile.objects.create(
            user=senior_user, organization=self.o1, role="manager"
        )
        manager_user = User.objects.create_user(
            "line-manager", email="manager@example.com", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager", manager=senior
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        sheet = Timesheet.objects.create(
            organization=self.o1,
            employee=self.p1,
            period_start=date(2026, 10, 1),
            period_end=date(2026, 10, 31),
            expert_signature=SimpleUploadedFile(
                "expert.png",
                base64.b64decode(
                    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
                ),
                content_type="image/png",
            ),
        )
        entry = TimesheetEntry.objects.create(
            timesheet=sheet,
            date=date(2026, 10, 2),
            task_performed="Initial activity",
            hours=8,
            days_worked=1,
        )
        self.client.force_login(self.u1)
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            self.client.post(
                reverse("timesheet_action", args=[sheet.pk]), {"action": "submit"}
            )
        sheet.refresh_from_db()
        self.assertEqual(sheet.requested_approver, manager)
        self.assertEqual(sheet.request_task.priority, "medium")
        self.assertEqual(sheet.request_task.assigned_to, manager)
        self.client.force_login(manager_user)
        edited = self.client.post(
            reverse("timesheet_entry_edit", args=[sheet.pk, entry.pk]),
            {
                "date": "2026-10-03",
                "task": "",
                "task_performed": "Manager corrected activity",
                "hours": "7",
                "days_worked": "1",
                "location": "Nairobi",
                "description": "",
                "notes": "",
            },
        )
        self.assertRedirects(edited, reverse("timesheet_detail", args=[sheet.pk]))
        entry.refresh_from_db()
        self.assertEqual(entry.task_performed, "Manager corrected activity")
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            moved = self.client.post(
                reverse("timesheet_review", args=[sheet.pk]),
                {"decision": "escalate", "notes": "Senior review required."},
            )
        self.assertRedirects(moved, reverse("timesheet_detail", args=[sheet.pk]))
        sheet.refresh_from_db()
        sheet.request_task.refresh_from_db()
        self.assertEqual(sheet.requested_approver, senior)
        self.assertEqual(sheet.request_task.assigned_to, senior)
        self.assertEqual(sheet.status, "submitted")
        self.client.force_login(senior_user)
        self.client.post(
            reverse("timesheet_review", args=[sheet.pk]),
            {"decision": "approved", "notes": "Approved."},
        )
        sheet.refresh_from_db()
        sheet.request_task.refresh_from_db()
        self.assertEqual(sheet.status, "approved")
        self.assertEqual(sheet.request_task.status, "completed")

    def test_manager_requests_timesheet_and_employee_hierarchy_is_available(self):
        manager_user = User.objects.create_user(
            "request-manager", email="manager@example.com", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.u1.email = "employee@example.com"
        self.u1.save(update_fields=["email"])
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        self.client.force_login(manager_user)
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(
                reverse("timesheet_request"),
                {
                    "employee": self.p1.pk,
                    "month": "11",
                    "year": "2026",
                    "due_date": "2026-11-05",
                    "instructions": "Include field work.",
                },
            )
        self.assertRedirects(response, reverse("timesheets"))
        sheet = Timesheet.objects.get(employee=self.p1, period_start=date(2026, 11, 1))
        self.assertEqual(sheet.requested_approver, manager)
        self.assertEqual(sheet.request_task.priority, "medium")
        self.assertEqual(sheet.request_task.assigned_to, self.p1)
        self.assertEqual(mail.outbox[-1].to, ["employee@example.com"])
        hierarchy = self.client.get(reverse("employees"))
        self.assertContains(hierarchy, "Hierarchy tree")
        self.assertContains(hierarchy, "Reports up")

    def test_senior_manager_can_request_timesheet_from_indirect_report(self):
        senior_user = User.objects.create_user("request-senior", password="testpass123")
        senior = Profile.objects.create(
            user=senior_user, organization=self.o1, role="manager"
        )
        manager_user = User.objects.create_user(
            "request-middle", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager", manager=senior
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        self.client.force_login(senior_user)
        employee_page = self.client.get(reverse("employees"))
        self.assertContains(
            employee_page, f'{reverse("timesheet_request")}?employee={self.p1.pk}'
        )
        request_page = self.client.get(
            f'{reverse("timesheet_request")}?employee={self.p1.pk}'
        )
        self.assertEqual(
            request_page.context["form"].fields["employee"].initial, [str(self.p1.pk)]
        )
        response = self.client.post(
            reverse("timesheet_request"),
            {
                "employee": self.p1.pk,
                "month": "12",
                "year": "2026",
                "due_date": "2026-12-04",
                "instructions": "Submit the monthly record.",
            },
        )
        self.assertRedirects(response, reverse("timesheets"))
        sheet = Timesheet.objects.get(employee=self.p1, period_start=date(2026, 12, 1))
        self.assertEqual(sheet.requested_approver, senior)
        self.assertEqual(sheet.request_task.created_by, senior_user)

    def test_manager_can_request_multiple_months_from_multiple_employees(self):
        manager_user = User.objects.create_user("batch-manager", password="testpass123")
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        second_user = User.objects.create_user(
            "batch-employee", email="batch@example.com", password="testpass123"
        )
        second = Profile.objects.create(
            user=second_user, organization=self.o1, role="employee", manager=manager
        )
        self.u1.email = "first@example.com"
        self.u1.save(update_fields=["email"])
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        self.client.force_login(manager_user)
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(
                reverse("timesheet_request"),
                {
                    "employee": [self.p1.pk, second.pk],
                    "month": ["10", "11"],
                    "year": "2026",
                    "due_date": "2026-10-05",
                    "instructions": "Complete both months.",
                },
            )
        self.assertRedirects(response, reverse("timesheets"))
        sheets = Timesheet.objects.filter(
            employee__in=[self.p1, second],
            period_start__year=2026,
            period_start__month__in=[10, 11],
        )
        self.assertEqual(sheets.count(), 4)
        self.assertEqual(
            Task.objects.filter(
                created_by=manager_user,
                title__startswith="Complete ",
                assigned_to__in=[self.p1, second],
            ).count(),
            4,
        )
        self.assertEqual(len(mail.outbox), 4)

    def test_item_request_routes_up_hierarchy_and_exports_all_formats(self):
        senior_user = User.objects.create_user(
            "supply-senior", email="senior@example.com", password="testpass123"
        )
        senior = Profile.objects.create(
            user=senior_user, organization=self.o1, role="manager"
        )
        manager_user = User.objects.create_user(
            "supply-manager", email="manager@example.com", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager", manager=senior
        )
        self.u1.email = "requester@example.com"
        self.u1.save(update_fields=["email"])
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        self.client.force_login(self.u1)
        data = {
            "purpose": "Office kitchen and cleaning supplies",
            "needed_by": "2026-09-10",
            "delivery_location": "Nairobi office",
            "department": "",
            "lines-TOTAL_FORMS": "2",
            "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "1",
            "lines-MAX_NUM_FORMS": "1000",
            "lines-0-item": "Sugar",
            "lines-0-quantity": "2 kg",
            "lines-0-estimated_cost": "305",
            "lines-0-source_link": "https://example.com/sugar",
            "lines-0-notes": "",
            "lines-1-item": "Tissues",
            "lines-1-quantity": "10 rolls",
            "lines-1-estimated_cost": "299",
            "lines-1-source_link": "https://example.com/tissues",
            "lines-1-notes": "Soft tissue",
        }
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(reverse("item_request_create"), data)
        obj = ItemRequest.objects.get(requested_by=self.u1)
        self.assertRedirects(response, reverse("item_request_detail", args=[obj.pk]))
        self.assertEqual(obj.total_estimated_cost, Decimal("604"))
        self.assertEqual(obj.requested_approver, manager)
        self.assertEqual(obj.approval_task.priority, "medium")
        self.assertEqual(mail.outbox[-1].to, ["manager@example.com"])
        for fmt, content_type in [
            ("pdf", "application/pdf"),
            (
                "docx",
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ),
            (
                "xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        ]:
            exported = self.client.get(
                reverse("item_request_export", args=[obj.pk, fmt])
            )
            self.assertEqual(exported.status_code, 200)
            self.assertEqual(exported["Content-Type"], content_type)
            self.assertGreater(len(exported.content), 500)
        from openpyxl import load_workbook

        workbook = load_workbook(
            BytesIO(
                self.client.get(
                    reverse("item_request_export", args=[obj.pk, "xlsx"])
                ).content
            ),
            data_only=False,
        )
        sheet = workbook["Consumable supplies"]
        self.assertEqual(sheet["A12"].value, "Sugar")
        self.assertEqual(sheet["C14"].value, "=SUM(C12:C13)")
        self.client.force_login(manager_user)
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            self.client.post(
                reverse("item_request_action", args=[obj.pk, "escalate"]),
                {"notes": "Please confirm budget."},
            )
        obj.refresh_from_db()
        obj.approval_task.refresh_from_db()
        self.assertEqual(obj.requested_approver, senior)
        self.assertEqual(obj.approval_task.assigned_to, senior)
        self.client.force_login(senior_user)
        self.client.post(
            reverse("item_request_action", args=[obj.pk, "approve"]),
            {"notes": "Approved."},
        )
        obj.refresh_from_db()
        obj.approval_task.refresh_from_db()
        self.assertEqual(obj.status, "approved")
        self.assertEqual(obj.approval_task.status, "completed")

    def test_rejected_item_request_completes_its_review_task(self):
        manager_user = User.objects.create_user(
            "rejecting-manager", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        task = Task.objects.create(
            organization=self.o1,
            title="Review supplies",
            assigned_to=manager,
            created_by=self.u1,
            priority="medium",
            status="assigned",
            due_date=timezone.localdate() + timedelta(days=2),
        )
        obj = ItemRequest.objects.create(
            organization=self.o1,
            number="IR-2026-0099",
            requested_by=self.u1,
            purpose="Office supplies",
            needed_by=timezone.localdate() + timedelta(days=7),
            requested_approver=manager,
            approval_task=task,
            status="submitted",
        )
        self.client.force_login(manager_user)
        response = self.client.post(
            reverse("item_request_action", args=[obj.pk, "reject"]),
            {"notes": "Revise the quantities."},
        )
        self.assertRedirects(response, reverse("item_request_detail", args=[obj.pk]))
        obj.refresh_from_db()
        task.refresh_from_db()
        self.assertEqual(obj.status, "rejected")
        self.assertEqual(task.status, "completed")

    def test_timesheet_csv_export_is_not_available(self):
        sheet = Timesheet.objects.create(
            organization=self.o1,
            employee=self.p1,
            period_start=date(2026, 12, 1),
            period_end=date(2026, 12, 31),
        )
        self.client.force_login(self.u1)
        self.assertEqual(
            self.client.get(
                reverse("timesheet_export", args=[sheet.pk, "csv"])
            ).status_code,
            404,
        )

    def test_employee_cannot_change_a_task_they_only_created(self):
        colleague_user = User.objects.create_user(
            "task-colleague", password="testpass123"
        )
        colleague = Profile.objects.create(
            user=colleague_user, organization=self.o1, role="employee"
        )
        task = Task.objects.create(
            organization=self.o1,
            title="Colleague task",
            assigned_to=colleague,
            created_by=self.u1,
            status="assigned",
            due_date=timezone.localdate() + timedelta(days=1),
        )
        self.client.force_login(self.u1)
        response = self.client.post(
            reverse("task_status", args=[task.pk]), {"status": "in_progress"}
        )
        self.assertRedirects(response, reverse("task_detail", args=[task.pk]))
        task.refresh_from_db()
        self.assertEqual(task.status, "assigned")

    def test_direct_chat_is_reused_and_is_tenant_scoped(self):
        colleague_user = User.objects.create_user("colleague", password="testpass123")
        colleague = Profile.objects.create(
            user=colleague_user, organization=self.o1, role="employee"
        )
        self.client.force_login(self.u1)
        first = self.client.post(
            reverse("direct_chat_create"), {"person": colleague.pk}
        )
        thread = ChatThread.objects.get(kind="direct", organization=self.o1)
        self.assertRedirects(first, reverse("chat_detail", args=[thread.pk]))
        self.client.post(reverse("direct_chat_create"), {"person": colleague.pk})
        self.assertEqual(
            ChatThread.objects.filter(kind="direct", organization=self.o1).count(), 1
        )
        self.assertEqual(thread.memberships.count(), 2)
        self.client.force_login(self.u2)
        self.assertEqual(
            self.client.get(reverse("chat_detail", args=[thread.pk])).status_code, 404
        )

    def test_chat_page_separates_private_and_group_threads_into_tabs(self):
        colleague_user = User.objects.create_user(
            "chat-colleague", password="testpass123"
        )
        colleague = Profile.objects.create(
            user=colleague_user, organization=self.o1, role="employee"
        )
        direct = ChatThread.objects.create(
            organization=self.o1,
            kind="direct",
            direct_key=f"{self.p1.pk}-{colleague.pk}",
            created_by=self.u1,
        )
        ChatMembership.objects.create(thread=direct, profile=self.p1, added_by=self.u1)
        ChatMembership.objects.create(
            thread=direct, profile=colleague, added_by=self.u1
        )
        group = ChatThread.objects.create(
            organization=self.o1,
            kind="group",
            name="Operations channel",
            created_by=self.u1,
        )
        ChatMembership.objects.create(
            thread=group, profile=self.p1, added_by=self.u1, is_admin=True
        )
        self.client.force_login(self.u1)
        response = self.client.get(reverse("chats"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-chat-tab="private"')
        self.assertContains(response, 'data-chat-tab="groups"')
        self.assertContains(response, 'data-chat-panel="private"')
        self.assertContains(response, 'data-chat-panel="groups"')
        self.assertEqual(
            [thread.pk for thread in response.context["direct_threads"]], [direct.pk]
        )
        self.assertEqual(
            [thread.pk for thread in response.context["group_threads"]], [group.pk]
        )

    def test_group_invitation_acceptance_and_member_only_messages(self):
        colleague_user = User.objects.create_user("colleague", password="testpass123")
        colleague = Profile.objects.create(
            user=colleague_user, organization=self.o1, role="employee"
        )
        self.client.force_login(self.u1)
        response = self.client.post(
            reverse("group_chat_create"),
            {"name": "Finance team", "invitees": [colleague.pk]},
        )
        thread = ChatThread.objects.get(name="Finance team")
        self.assertRedirects(response, reverse("chat_detail", args=[thread.pk]))
        invitation = ChatInvitation.objects.get(
            thread=thread, invitee=colleague, status="pending"
        )
        self.assertFalse(
            ChatMembership.objects.filter(thread=thread, profile=colleague).exists()
        )

        self.client.force_login(colleague_user)
        self.client.post(
            reverse("chat_invitation_action", args=[invitation.pk, "accept"])
        )
        self.assertTrue(
            ChatMembership.objects.filter(thread=thread, profile=colleague).exists()
        )
        sent = self.client.post(
            reverse("chat_detail", args=[thread.pk]),
            {"body": "Hello group"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(sent.status_code, 200)
        message = ChatMessage.objects.get(thread=thread)
        feed = self.client.get(
            reverse("chat_message_feed", args=[thread.pk]), {"after": 0}
        )
        self.assertEqual(feed.json()["messages"][0]["id"], message.pk)

        self.client.force_login(self.u2)
        self.assertEqual(
            self.client.get(reverse("chat_message_feed", args=[thread.pk])).status_code,
            404,
        )

    def test_authenticated_activity_updates_online_presence(self):
        self.client.force_login(self.u1)
        self.client.get(reverse("chats"))
        presence = ChatPresence.objects.get(profile=self.p1)
        self.assertTrue(presence.is_online)

    def test_manager_assigns_leave_and_employee_requests_working_days(self):
        manager_user = User.objects.create_user(
            "leave-manager", email="manager@example.com", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        self.client.force_login(manager_user)
        response = self.client.post(
            reverse("leave_allocation_create"),
            {
                "employee": self.p1.pk,
                "leave_type": "annual",
                "year": 2026,
                "allocated_days": 10,
                "notes": "Annual allocation",
            },
        )
        self.assertRedirects(response, reverse("leave_dashboard"))
        allocation = LeaveAllocation.objects.get(
            employee=self.p1, leave_type="annual", year=2026
        )
        self.assertEqual(allocation.allocated_days, 10)

        self.u1.email = "employee@example.com"
        self.u1.save(update_fields=["email"])
        self.client.force_login(self.u1)
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(
                reverse("leave_request_create"),
                {
                    "leave_type": "annual",
                    "start_date": "2026-09-07",
                    "end_date": "2026-09-11",
                    "reason": "Family trip",
                },
            )
        self.assertRedirects(response, reverse("leave_dashboard"))
        leave = LeaveRequest.objects.get(employee=self.p1)
        self.assertEqual(leave.days, 5)
        self.assertEqual(leave.requested_approver, manager)
        self.assertEqual(mail.outbox[-1].to, ["manager@example.com"])

        self.client.force_login(manager_user)
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(
                reverse("leave_request_review", args=[leave.pk]),
                {"decision": "approved", "message": "Enjoy your leave."},
            )
        self.assertRedirects(response, reverse("leave_dashboard"))
        leave.refresh_from_db()
        self.assertEqual(leave.status, "approved")
        self.assertEqual(leave.review_message, "Enjoy your leave.")
        self.assertEqual(leave.reviewed_by, manager_user)
        self.assertEqual(mail.outbox[-1].to, ["employee@example.com"])

    def test_manager_can_reject_leave_without_message_and_cannot_manage_other_tenant(
        self,
    ):
        manager_user = User.objects.create_user("leave-manager", password="testpass123")
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        allocation = LeaveAllocation.objects.create(
            organization=self.o1,
            employee=self.p1,
            leave_type="sick",
            year=2026,
            allocated_days=5,
            assigned_by=manager_user,
        )
        leave = LeaveRequest.objects.create(
            organization=self.o1,
            employee=self.p1,
            requested_approver=manager,
            leave_type="sick",
            start_date=date(2026, 9, 7),
            end_date=date(2026, 9, 8),
            days=2,
            reason="Recovery",
        )
        self.client.force_login(manager_user)
        response = self.client.post(
            reverse("leave_request_review", args=[leave.pk]),
            {"decision": "rejected", "message": ""},
        )
        self.assertRedirects(response, reverse("leave_dashboard"))
        leave.refresh_from_db()
        self.assertEqual(leave.status, "rejected")
        self.assertEqual(leave.review_message, "")
        response = self.client.post(
            reverse("leave_allocation_create"),
            {
                "employee": self.p2.pk,
                "leave_type": "annual",
                "year": 2026,
                "allocated_days": 20,
                "notes": "",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            LeaveAllocation.objects.filter(
                organization=self.o2, employee=self.p2
            ).exists()
        )

    def test_leave_request_can_move_up_the_management_chain(self):
        senior_user = User.objects.create_user(
            "leave-senior", email="senior@example.com", password="testpass123"
        )
        senior = Profile.objects.create(
            user=senior_user, organization=self.o1, role="manager"
        )
        manager_user = User.objects.create_user(
            "leave-middle", email="middle@example.com", password="testpass123"
        )
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager", manager=senior
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        LeaveAllocation.objects.create(
            organization=self.o1,
            employee=self.p1,
            leave_type="annual",
            year=2026,
            allocated_days=10,
            assigned_by=manager_user,
        )
        leave = LeaveRequest.objects.create(
            organization=self.o1,
            employee=self.p1,
            requested_approver=manager,
            leave_type="annual",
            start_date=date(2026, 9, 7),
            end_date=date(2026, 9, 9),
            days=3,
            reason="Family trip",
        )
        self.client.force_login(manager_user)
        popup = self.client.get(reverse("leave_request_review", args=[leave.pk]))
        self.assertContains(popup, "Move to my manager")
        with self.settings(
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"
        ):
            response = self.client.post(
                reverse("leave_request_review", args=[leave.pk]),
                {"decision": "escalate", "message": "Please review."},
            )
        self.assertRedirects(response, reverse("leave_dashboard"))
        leave.refresh_from_db()
        self.assertEqual(leave.status, "pending")
        self.assertEqual(leave.requested_approver, senior)
        self.assertEqual(mail.outbox[-1].to, ["senior@example.com"])
        self.client.force_login(senior_user)
        popup = self.client.get(reverse("leave_request_review", args=[leave.pk]))
        self.assertNotContains(popup, "Move to my manager")
        self.client.post(
            reverse("leave_request_review", args=[leave.pk]),
            {"decision": "approved", "message": "Approved."},
        )
        leave.refresh_from_db()
        self.assertEqual(leave.status, "approved")
        self.assertEqual(leave.reviewed_by, senior_user)

    def test_leave_request_cannot_exceed_balance_or_overlap(self):
        manager_user = User.objects.create_user("leave-manager", password="testpass123")
        manager = Profile.objects.create(
            user=manager_user, organization=self.o1, role="manager"
        )
        self.p1.manager = manager
        self.p1.save(update_fields=["manager"])
        LeaveAllocation.objects.create(
            organization=self.o1,
            employee=self.p1,
            leave_type="annual",
            year=2026,
            allocated_days=5,
            assigned_by=manager_user,
        )
        LeaveRequest.objects.create(
            organization=self.o1,
            employee=self.p1,
            requested_approver=manager,
            leave_type="annual",
            start_date=date(2026, 9, 7),
            end_date=date(2026, 9, 9),
            days=3,
            reason="Existing",
            status="pending",
        )
        self.client.force_login(self.u1)
        overlap = self.client.post(
            reverse("leave_request_create"),
            {
                "leave_type": "annual",
                "start_date": "2026-09-09",
                "end_date": "2026-09-10",
                "reason": "Overlap",
            },
        )
        self.assertEqual(overlap.status_code, 200)
        self.assertContains(
            overlap, "overlap another pending or approved leave request"
        )
        too_many = self.client.post(
            reverse("leave_request_create"),
            {
                "leave_type": "annual",
                "start_date": "2026-09-14",
                "end_date": "2026-09-16",
                "reason": "Too many",
            },
        )
        self.assertEqual(too_many.status_code, 200)
        self.assertContains(too_many, "only 2 are available")
        self.assertEqual(LeaveRequest.objects.filter(employee=self.p1).count(), 1)
